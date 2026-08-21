# -*- coding: utf-8 -*-
"""주소 이미지 수백 장을 한 번에 돌려 판독 정확도를 실측한다.

테스터가 하루 20건씩 손으로 돌리는 표본으로는 "챗봇에 주문을 맡겨도 되는가"를
판단할 수 없다. 몇 달치 실제 대화에서 모은 이미지를 통째로 돌려 숫자를 낸다.

측정의 축은 두 개다.
  1. 우편번호 종류 수     주소를 실재하는 한 곳으로 좁혔는가
  2. 이미지에 적힌 우편번호와의 대조   정답이 확실한 건의 적중률

정확도를 가르는 것은 글자 판독이 아니라 base/detail 분리다. 실측에서 판독은
대부분 맞는데 base 에 상호·건물설명·배송메모가 섞여 0건이 났다.
  첨단중앙로 136(첨단점 우리은행/농협3층)  → 0건
  첨단중앙로 136                          → 62276
그래서 프롬프트의 절반이 "무엇을 base 에 넣지 않는가" 규칙이다.

  python batch_address.py <이미지폴더> [--out 결과.xlsx] [--model 모델명]
                          [--limit N] [--workers 4] [--repeat 2]

중단해도 된다. 처리한 건은 진행.jsonl 에 즉시 적히고 다시 돌리면 건너뛴다.
엑셀은 매 실행 끝에 그 jsonl 전체를 다시 그려 만든다.

※ 실제 고객 정보다. 이미지도 결과도 저장소에 올리지 않는다(.gitignore 확인).
"""
import argparse
import io
import json
import os
import random
import re
import shutil
import sys
import threading
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from lib import images as IMG      # noqa: E402
from lib import juso as JUSO       # noqa: E402
from lib import llm as LLM         # noqa: E402

# 콘솔이 cp949 라도 한글 요약이 깨지지 않게 한다
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BASE = os.path.dirname(os.path.abspath(__file__))

EXTS = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
        ".webp": "image/webp", ".gif": "image/gif", ".bmp": "image/bmp"}

# 비용 표시용 환율. 청구액이 아니라 규모 감각을 주기 위한 값이다.
USD_KRW = 1380

VERDICTS = ("성공", "불완전", "실패", "원본불량")

# 주소 API 가 연달아 이만큼 실패하면 멈춘다.
# 일일 호출 한도를 넘기면 주소 API 만 죽고 Gemini 는 계속 정상 호출된다.
# 즉 돈은 계속 나가는데 채점은 전부 오류로 쌓인다. 13,000장이면 그 낭비가 크다.
# 한 건 실패는 일시적 네트워크일 수 있으니 연속으로 셀 때만 멈춘다.
ERR_LIMIT = 20
_ERR = {"streak": 0, "stop": False, "why": ""}
_ERR_LOCK = threading.Lock()


# ------------------------------------------------------------------ 설정 읽기
def secret(key):
    """secrets.toml 에서 키를 읽는다.

    lib.sheets.secret 은 streamlit 런타임 기준으로 경로를 잡아서, 다른 폴더에서
    이 스크립트를 실행하면 조용히 빈 값을 준다. 키가 없으면 500장이 전부 오류로
    끝나므로 프로젝트 폴더의 파일을 직접 읽는 경로를 뒤에 둔다."""
    try:
        from lib import sheets as SH
        v = SH.secret(key)
        if v:
            return v
    except Exception:
        pass
    try:
        import tomllib
        with open(os.path.join(BASE, ".streamlit", "secrets.toml"), "rb") as f:
            return tomllib.load(f).get(key)
    except Exception:
        return None


def default_model():
    m = secret("MODELS")
    if isinstance(m, (list, tuple)) and m:
        return str(m[0])
    return "gemini-3.5-flash"


# ------------------------------------------------------------------ 프롬프트
SYSTEM = """\
이미지에서 택배 받는 사람의 정보를 읽어 JSON 하나만 출력한다. 설명은 쓰지 않는다.

{"수령인": null, "전화": null, "주소_base": null, "주소_detail": null,
 "이미지내_우편번호": null, "본_전화번호_전부": [], "본_이름후보_전부": [],
 "가려진_항목": [], "원본불량_사유": null, "이미지_문제": []}

[1] 받는 사람(To)만 읽는다
송장·화면 캡처에는 보내는 사람(From)과 받는 사람이 같이 찍혀 있는 일이 많다.
이름 두 개, 전화 두 개가 보이면 어느 쪽이 받는 사람인지 배치와 표시로 판단한다.
발송인을 수령인으로 읽으면 물건이 되돌아간다.

[2] 주소_base 와 주소_detail 을 반드시 나눈다
base   시·도부터 도로명·건물번호까지, 또는 지번 주소까지만. 그 뒤는 전부 잘라낸다.
detail base 에 넣지 않은 나머지 전부.
       동·호수, 아파트·건물 이름, 상호·법인명, 농장명, 괄호 안 건물 설명,
       출입 코드, 배송 메모(문 앞, 경비실 등).

주소 검색 API 는 base 에 주소가 아닌 것이 한 조각이라도 섞이면 0건을 낸다.
아래는 실제로 0건이 났던 것들이다. 오른쪽처럼 갈라야 한다.

  첨단중앙로 136(첨단점 우리은행/농협3층)
    base "첨단중앙로 136"            detail "첨단점 우리은행/농협3층"
  금암1길 50 농업회사법인 유한회사 도원바이오
    base "금암1길 50"                detail "농업회사법인 유한회사 도원바이오"
  갯벌체험로944번길10. 삼해수산. 태오
    base "갯벌체험로944번길 10"       detail "삼해수산, 태오"
  육동로 727-1 버섯 농장
    base "육동로 727-1"              detail "버섯 농장"
  단북1길 55 문앞에 놓고 사진찍어보내주세요
    base "단북1길 55"                detail "문앞에 놓고 사진찍어보내주세요"

시·군이 안 적혀 있어도 도로명과 번호만으로 검색되므로 없는 시·군을 지어내지 않는다.
띄어쓰기는 없어도 되고 지번이어도 된다. base 에서 우편번호는 빼고 아래 칸에 적는다.

[3] 가려졌거나 잘린 값은 비워 둔다
별표(dar**), 검은 칠, 빨간 칠, 모자이크로 가려졌거나 사진 밖으로 잘려나갔거나
너무 흐려 글자를 못 세는 값은 추측해 채우지 말고 null 로 둔다.
그런 항목만 "가려진_항목" 에 "수령인" "전화" "주소" 중 해당하는 것을 넣고,
"원본불량_사유" 에 가려짐·잘림·흐림 중 무엇인지 한 줄로 적는다.

"가려진_항목" 은 "이미지가 나빠서 못 읽은 것" 만이다. 애초에 적혀 있지 않은 정보는
가려진 것이 아니다. 이 둘을 섞으면 이미지 품질 통계가 통째로 틀어진다.
  이름 자리에 dar** 라고 찍혀 있다        → 가려짐 (넣는다)
  도로명판 사진이라 이름·전화가 없다       → 없는 것 (넣지 않는다)
  주소 윗줄이 화면 밖으로 잘렸다          → 가려짐 (넣는다)
애매하면 넣지 않는다.

[4] 후보는 고른 것과 별개로 전부 적는다
본_전화번호_전부  이미지에서 본 전화번호를 순서대로 전부. 유선번호도 포함한다.
본_이름후보_전부  이미지에서 본 사람 이름·상호를 전부. 발송인 이름도 포함한다.
나중에 사람이 "고른 것이 맞았는지" 보기 위한 목록이다.

[5] 이미지에 우편번호가 적혀 있으면
[59107] (26367) 05544 같은 5자리 숫자를 "이미지내_우편번호" 에 숫자만 넣는다.
없으면 null. 지어내지 않는다.

[6] 이 사진 자체의 문제를 분류한다
읽기를 어렵게 만든 요인을 "이미지_문제" 에 아래 낱말로만 넣는다. 여러 개면 여러 개.
문제가 없으면 ["정상"] 하나만 넣는다. 아래 목록에 없는 말은 쓰지 않는다.

  두사람정보   발송인과 수령인 정보가 한 장에 같이 있다
  가려짐       별표·검은칠·모자이크로 지워진 값이 있다
  잘림         글자가 사진 밖으로 나가 일부만 보인다
  흐림         초점이 안 맞거나 손떨림으로 번졌다
  기울어짐     비스듬히 찍혀 글자가 휘었다
  반사         화면을 다시 찍어 빛반사·물결무늬가 있다
  글씨작음     해상도에 비해 글자가 너무 작다
  손글씨       손으로 쓴 글씨가 섞여 있다
  주소불완전   주소가 도로명·번호까지 다 적혀 있지 않다
  정상         위 어느 것도 해당하지 않는다

이건 판독 성패와 별개로 "어떤 사진이 문제를 일으키는가" 를 모으기 위한 것이다.
잘 읽혔더라도 해당하면 넣는다.

이름은 Nuy, 태오, Yada Narktung, 아논 코 신차이 처럼 제각각이다. 한국식으로 고치지 않고
적힌 그대로 옮긴다.
"""


def extract(path, model, api_key):
    """이미지 한 장을 읽어 (추출 dict, 사용량, 지연초, 오류문자열)."""
    t0 = time.time()
    try:
        with open(path, "rb") as f:
            raw = f.read()
    except Exception as e:
        return None, {}, 0.0, "파일 읽기 실패: %s" % e

    mime = EXTS.get(os.path.splitext(path)[1].lower(), "image/jpeg")
    data, mime, _ = IMG.prepare(raw, mime)

    try:
        # attempts=1 로 부른다.
        # lib.llm.call 은 429 를 "일시적 과부하" 로 보고 3회까지 다시 부르는데,
        # 하루 요청 한도 초과는 재시도해도 절대 풀리지 않는다. 실제로 이것 때문에
        # 남은 이미지 10,426장이 각각 3번씩, 총 3만 건이 넘는 요청을 태워
        # 그날 한도를 순식간에 날렸다. 여기서는 한 번만 부르고,
        # 오류가 이어지면 아래 감시가 즉시 멈춘다.
        out, text, usage = LLM.call(api_key, model, SYSTEM, "이미지를 읽어라.",
                                    [{"bytes": data, "mime": mime}], attempts=1,
                                    schema=None)
    except Exception as e:
        return None, {}, time.time() - t0, "%s: %s" % (type(e).__name__, e)

    if not isinstance(out, dict):
        # 판독을 틀린 것이 아니라 응답을 못 받은 것이다. 실패로 세면 안 된다.
        return None, usage, time.time() - t0, "JSON 파싱 실패 (%d자)" % len(text or "")

    return out, usage, time.time() - t0, ""


# ------------------------------------------------------------------ 채점
_DIGITS = re.compile(r"\D")
# 이름에 섞이면 주소를 이름으로 읽은 것으로 보는 글자.
# 흔한 한국 이름에 들어가는 글자(동·리 등)까지 넣으면 멀쩡한 이름이 오염으로 잡히므로
# 지시받은 다섯 글자만 본다. 어차피 사람이 표본 검수로 확인할 지표다.
_ADDR_CHARS = ("시", "군", "구", "로", "길")


def phone_form(value):
    """형식만 본다. 여러 번호 중 옳은 것을 골랐는지는 자동으로 알 수 없다."""
    d = _DIGITS.sub("", str(value or ""))
    if not d:
        return "빈값"
    if d.startswith("010") and 10 <= len(d) <= 11:
        return "유효"
    return "실패"


def name_flag(value, masked):
    """자동으로는 오염만 잡는다. 이름 자체가 맞는지는 표본 검수로 본다."""
    s = str(value or "").strip()
    if not s:
        return "빈값(가려짐)" if masked else "빈값"
    if re.search(r"\d", s):
        return "숫자"
    if any(c in s for c in _ADDR_CHARS):
        return "주소어"
    return ""


def judge(base, zip_kinds, addr_masked, error):
    """주소 판정. 우편번호 종류 수로 가른다. 검색 건수(total)가 아니다.

    불완전을 실패에 묶지 않는다. 불완전은 상담원이 한 번 물어보면 끝나고,
    실패는 판독 자체가 틀린 것이라 대응이 완전히 다르다.
    원본불량도 실패에 묶지 않는다. 가려진 이름이나 잘린 상단을 못 읽는 것은
    정상이며, 실패로 세면 정확도가 실제보다 낮게 나온다.

    다만 주소가 가려졌어도 우편번호가 한 종류로 나왔다면 그것은 성공이다.
    원본불량은 '이미지 때문에 못 읽은 것'에만 붙인다."""
    if error:
        return "오류"
    if zip_kinds == 1:
        return "성공"
    if zip_kinds >= 2:
        return "불완전"
    if addr_masked or not (base or "").strip():
        return "원본불량" if addr_masked else "실패"
    return "실패"


def process(path, root, model, api_key, run):
    """이미지 한 장을 끝까지 처리해 기록 한 줄을 만든다."""
    rel = os.path.relpath(path, root)
    # 날짜별 폴더로 넣는 경우가 많다. 어느 폴더에서 온 장인지 남겨두면
    # 엑셀에서 폴더로 걸러 "이 날짜 묶음만 유독 실패가 많다"를 볼 수 있다.
    sub = os.path.dirname(rel)
    rec = {"경로": os.path.abspath(path), "파일명": os.path.basename(path),
           "폴더": sub.replace("\\", "/") or "(최상위)",
           "상대경로": rel, "회차": run, "모델": model}

    out, usage, sec, err = extract(path, model, api_key)
    rec["지연초"] = round(sec, 2)
    rec["토큰"] = int((usage or {}).get("input") or 0) + int((usage or {}).get("output") or 0)
    rec["입력토큰"] = int((usage or {}).get("input") or 0)
    rec["출력토큰"] = int((usage or {}).get("output") or 0)
    out = out or {}

    masked = [str(x) for x in (out.get("가려진_항목") or [])]
    base = str(out.get("주소_base") or "").strip()

    rec.update({
        "수령인": str(out.get("수령인") or "").strip(),
        "전화": str(out.get("전화") or "").strip(),
        "주소_base": base,
        "주소_detail": str(out.get("주소_detail") or "").strip(),
        "이미지내_우편번호": _DIGITS.sub("", str(out.get("이미지내_우편번호") or "")),
        "본_이름후보_전부": [str(x) for x in (out.get("본_이름후보_전부") or [])],
        "본_전화번호_전부": [str(x) for x in (out.get("본_전화번호_전부") or [])],
        "가려진_항목": masked,
        "이미지_문제": [str(x) for x in (out.get("이미지_문제") or [])],
        "원본불량_사유": str(out.get("원본불량_사유") or "").strip(),
        "오류": err,
    })

    # base 만 넣는다. detail 을 붙이면 상호·메모 때문에 0건이 난다
    zips, total, zipno, road, used = [], 0, "", "", ""
    if base and not err:
        confm = secret("JUSO_CONFM_KEY")
        used = base
        r = JUSO.search(base, confm)
        if r.get("error"):
            rec["오류"] = "주소API: %s" % r["error"]

        # 0건이면 시·도를 떼고 한 번만 더 본다. 고객의 시·도 오기를 구제한다
        if not (r.get("zips") or []) and not r.get("error"):
            alt = without_sido(base)
            if alt:
                r2 = JUSO.search(alt, confm)
                if r2.get("zips"):
                    r, used = r2, alt
                    rec["시도교정"] = "Y"

        zips = r.get("zips") or []
        total = int(r.get("total") or 0)
        zipno = r.get("zipno") or ""
        road = r.get("road_addr") or ""
    rec["주소_검색어"] = used

    # 실패가 쌓이면 멈춘다. 주소 API 뿐 아니라 LLM 쪽 한도도 여기서 잡는다.
    # 실제로 Gemini 하루 한도(10,000건)에 걸렸을 때 이 감시가 주소 API 만 보고 있어서
    # 남은 10,430장을 그대로 태웠다. 32분이 통째로 버려졌다.
    with _ERR_LOCK:
        if rec["오류"]:
            _ERR["streak"] += 1
            # 하루 한도는 기다리는 것 말고 방법이 없다. 즉시 멈춘다
            fatal = any(k in rec["오류"].lower()
                        for k in ("per_day", "perday", "resource_exhausted", "quota"))
            if (fatal or _ERR["streak"] >= ERR_LIMIT) and not _ERR["stop"]:
                _ERR["stop"] = True
                _ERR["why"] = rec["오류"][:300]
        else:
            _ERR["streak"] = 0

    printed = rec["이미지내_우편번호"]
    rec.update({
        "우편번호_종류수": len(zips),
        "우편번호": zipno,
        "도로명주소": road,
        "검색건수": total,
        "판정": judge(base, len(zips), "주소" in masked, rec["오류"]),
        "전화형식": phone_form(rec["전화"]),
        "이름오염": name_flag(rec["수령인"], "수령인" in masked),
        "우편번호일치": ("일치" if printed in zips else "불일치") if printed and zips else "",
    })
    return rec


# 고객이 직접 쓴 주소에는 시·도를 잘못 적은 것이 섞여 있다.
#   "충청북도 논산시 양촌면 중산1길 239"  (논산시는 충청남도다)
#   "경상북도 의령군 화정면 상일상이들8길 24" (의령군은 경상남도다)
# 둘 다 실물 이미지를 확인했다. AI 판독은 정확했고 고객이 틀리게 썼다.
# 광역시·도만 떼면 API 가 정확히 찾는다. 시·군·구는 반드시 남긴다 —
# 거기까지 떼면 같은 이름의 다른 동네가 1종으로 잡혀 틀린 우편번호를 성공으로 센다.
_SIDO = re.compile(
    r"^(서울특별시|서울시|서울|부산광역시|부산|대구광역시|대구|인천광역시|인천"
    r"|광주광역시|대전광역시|대전|울산광역시|울산|세종특별자치시|세종|경기도|경기"
    r"|강원특별자치도|강원도|강원|충청북도|충북|충청남도|충남|전북특별자치도|전라북도|전북"
    r"|전라남도|전남|경상북도|경북|경상남도|경남|제주특별자치도|제주도|제주)\s+")


def without_sido(base):
    """시·도를 뗀 검색어. 뗄 수 없거나 떼면 위험하면 None."""
    s = _SIDO.sub("", (base or "").strip(), count=1)
    if s == (base or "").strip():
        return None
    return s if re.search(r"(시|군|구)\s", s + " ") else None


# ------------------------------------------------------------------ 진행 기록
class Store:
    """처리한 건을 즉시 jsonl 에 적는다.

    수백 장이면 중간에 끊긴다. 끝까지 돌아야만 결과가 남는 구조면 그때마다
    처음부터 다시 돌려야 하고, 그 돈과 시간이 그대로 버려진다."""

    def __init__(self, path):
        self.path = path
        self.lock = threading.Lock()
        self.rows = []
        self.done = set()
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        r = json.loads(line)
                    except Exception:
                        continue
                    self.rows.append(r)
                    # 오류 행은 처리한 것으로 치지 않는다.
                    # 한도 초과·네트워크 실패로 남은 행까지 "완료"로 세면,
                    # 다음 날 다시 돌려도 그 이미지는 영영 건너뛴다.
                    if not r.get("오류"):
                        self.done.add((r.get("상대경로"), r.get("회차")))

    def add(self, rec):
        with self.lock:
            self.rows.append(rec)
            self.done.add((rec["상대경로"], rec["회차"]))
            with open(self.path, "a", encoding="utf-8") as f:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")


# ------------------------------------------------------------------ 산출물
COLUMNS = ["원본파일경로", "폴더", "파일명", "수령인", "본_이름후보_전부", "전화",
           "본_전화번호_전부", "주소_base", "주소_detail", "이미지내_우편번호",
           "우편번호_종류수", "우편번호", "도로명주소", "판정", "전화형식",
           "이름오염", "지연초", "토큰", "모델",
           "우편번호일치", "검색건수", "주소_검색어", "시도교정", "이미지_문제", "가려진_항목", "원본불량_사유",
           "회차", "재현일치", "오류"]


def dedupe(rows):
    """같은 (파일, 회차) 가 두 번 있으면 나중 것을 쓴다.

    한도에 걸려 오류로 남은 행을 다음 날 다시 돌리면 같은 키가 또 들어온다.
    둘 다 집계하면 오류가 통계에 그대로 남아 성공률이 실제보다 낮게 나온다."""
    latest = {}
    for r in rows:
        latest[(r.get("상대경로"), r.get("회차"))] = r
    return list(latest.values())


def mark_repeat(rows):
    """같은 이미지를 두 번 돌려 답이 갈리면 운으로 맞은 것이다.

    LLM 은 매번 같은 답을 주지 않는다. 프로덕션에 맡길지 판단하려면
    '몇 % 맞았나' 만큼이나 '같은 입력에 같은 답을 주는가' 가 필요하다."""
    by_file = {}
    for r in rows:
        by_file.setdefault(r.get("상대경로"), []).append(r)

    for group in by_file.values():
        if len(group) < 2:
            for r in group:
                r["재현일치"] = ""
            continue
        keys = {(_DIGITS.sub("", r.get("전화") or ""),
                 (r.get("수령인") or "").strip(),
                 re.sub(r"\s+", "", r.get("주소_base") or ""),
                 r.get("판정")) for r in group}
        verdict = "일치" if len(keys) == 1 else "불일치"
        for r in group:
            r["재현일치"] = verdict
    return rows


def row_values(r):
    join = lambda v: " | ".join(str(x) for x in (v or []))   # noqa: E731
    return [r.get("경로", ""), r.get("폴더", ""), r.get("파일명", ""), r.get("수령인", ""),
            join(r.get("본_이름후보_전부")), r.get("전화", ""),
            join(r.get("본_전화번호_전부")), r.get("주소_base", ""),
            r.get("주소_detail", ""), r.get("이미지내_우편번호", ""),
            r.get("우편번호_종류수", 0), r.get("우편번호", ""), r.get("도로명주소", ""),
            r.get("판정", ""), r.get("전화형식", ""), r.get("이름오염", ""),
            r.get("지연초", 0), r.get("토큰", 0), r.get("모델", ""),
            r.get("우편번호일치", ""), r.get("검색건수", 0),
            r.get("주소_검색어", ""), r.get("시도교정", ""),
            join(r.get("이미지_문제")), join(r.get("가려진_항목")),
            r.get("원본불량_사유", ""),
            r.get("회차", 1), r.get("재현일치", ""), r.get("오류", "")]


def write_xlsx(rows, path):
    """원본 파일 경로 열은 클릭하면 그 이미지가 열리게 한다.

    이상한 행을 발견했을 때 바로 이미지를 봐야 하는데, 파일명만 들고
    수백 개 폴더에서 찾는 것은 그 자체로 검수를 포기하게 만든다."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "결과"
    ws.append(COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E8EAED")
        c.alignment = Alignment(vertical="center")

    for r in sorted(rows, key=lambda x: (x.get("상대경로", ""), x.get("회차", 1))):
        ws.append(row_values(r))
        cell = ws.cell(row=ws.max_row, column=1)
        try:
            cell.hyperlink = "file:///" + str(cell.value).replace("\\", "/")
            cell.font = Font(color="1155CC", underline="single")
        except Exception:
            pass

    widths = {"원본파일경로": 46, "폴더": 18, "파일명": 30, "수령인": 16, "본_이름후보_전부": 26,
              "전화": 15, "본_전화번호_전부": 26, "주소_base": 40, "주소_detail": 34,
              "도로명주소": 44, "주소_검색어": 34, "이미지_문제": 24, "원본불량_사유": 22, "오류": 24}
    for i, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 13)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    wb.save(path)


def copy_by_verdict(rows, outdir):
    """판정별 폴더로 복사한다. 이동이 아니라 복사다.

    이동하면 두 번째 실행 때 원본 폴더가 비어 재현이 안 된다.
    실패 폴더가 진짜 산출물이다. 모아놓고 눈으로 훑으면 실패 유형
    (흐림·기울어짐·잘림·가려짐·발송인 오독)이 보이고, 그것이 고객 안내 문구나
    프롬프트 수정으로 이어진다."""
    first = {}
    for r in sorted(rows, key=lambda x: x.get("회차", 1)):
        first.setdefault(r.get("상대경로"), r)

    n = 0
    for rel, r in first.items():
        src = r.get("경로")
        if not src or not os.path.exists(src):
            continue
        dst_dir = os.path.join(outdir, r.get("판정") or "오류")
        os.makedirs(dst_dir, exist_ok=True)
        # 하위 폴더가 여러 개면 파일명이 겹친다. 상대경로를 이름에 녹인다
        flat = str(rel).replace("\\", "_").replace("/", "_")
        # 판정이 바뀌면 이전 폴더의 복사본을 지운다.
        # 한도에 걸려 '오류' 로 분류됐다가 다시 돌려 '성공' 이 된 이미지가
        # 두 폴더에 동시에 남으면, 실패 폴더를 훑는 작업 자체가 못 믿을 것이 된다.
        for other in VERDICTS + ("오류",):
            if other == (r.get("판정") or "오류"):
                continue
            stale = os.path.join(outdir, other, flat)
            if os.path.exists(stale):
                try:
                    os.remove(stale)
                except Exception:
                    pass

        dst = os.path.join(dst_dir, flat)
        if os.path.exists(dst):
            continue
        try:
            shutil.copy2(src, dst)
            n += 1
        except Exception:
            pass
    return n


# ------------------------------------------------------------------ 요약
def pct(n, d):
    return "%5.1f%%" % (100.0 * n / d) if d else "    -  "


def summarize(rows):
    if not rows:
        print("기록이 없습니다.")
        return

    files = {r.get("상대경로") for r in rows}
    v = Counter(r.get("판정") for r in rows)
    n = len(rows)

    print()
    print("=" * 62)
    print("이미지 %d장 · 판독 %d건 (회차 포함)" % (len(files), n))
    print("=" * 62)

    print("[주소]  우편번호 종류 수로 판정")
    for k in VERDICTS:
        print("  %-8s %5d건  %s" % (k, v[k], pct(v[k], n)))
    if v["오류"]:
        print("  %-8s %5d건  %s   (API·파싱 실패. 판독 오류가 아니므로 따로 센다)"
              % ("오류", v["오류"], pct(v["오류"], n)))
    scored = n - v["오류"]
    if scored:
        print("  ─ 오류 제외 %d건 기준 성공 %s · 불완전 %s · 실패 %s"
              % (scored, pct(v["성공"], scored), pct(v["불완전"], scored),
                 pct(v["실패"], scored)))

    cross = [r for r in rows if r.get("우편번호일치")]
    hit = sum(1 for r in cross if r["우편번호일치"] == "일치")
    print("[교차검증]  이미지에 우편번호가 적혀 있던 건 — 정답이 확실한 표본")
    if cross:
        print("  %d건 중 %d건 일치  %s   ← 가장 신뢰도 높은 지표"
              % (len(cross), hit, pct(hit, len(cross))))
    else:
        print("  해당 없음")

    ph = Counter(r.get("전화형식") for r in rows)
    print("[전화]  형식만 자동 채점 (여러 번호 중 옳은 것을 골랐는지는 사람이 본다)")
    print("  유효 %d · 실패 %d · 빈값 %d   유효율 %s"
          % (ph["유효"], ph["실패"], ph["빈값"], pct(ph["유효"], n)))
    multi = sum(1 for r in rows if len(r.get("본_전화번호_전부") or []) > 1)
    print("  번호가 둘 이상 보인 건 %d  %s  ← 송장은 발송인·수령인이 같이 찍힌다"
          % (multi, pct(multi, n)))

    nm = Counter(r.get("이름오염") for r in rows)
    dirty = nm["숫자"] + nm["주소어"]
    print("[수령인]  자동으로는 오염만 잡는다 (정확도는 40~50건 표본 검수로 추정)")
    print("  오염 %d건 %s  (숫자 %d · 주소어 %d) · 빈값 %d · 빈값(가려짐) %d"
          % (dirty, pct(dirty, n), nm["숫자"], nm["주소어"], nm["빈값"], nm["빈값(가려짐)"]))

    rep = [r for r in rows if r.get("재현일치")]
    if rep:
        same = {r["상대경로"] for r in rep if r["재현일치"] == "일치"}
        tot = {r["상대경로"] for r in rep}
        print("[재현성]  같은 이미지를 다시 돌렸을 때 같은 답이 나온 비율")
        print("  %d장 중 %d장 일치  %s   ← 갈린 건은 운으로 맞은 것이다"
              % (len(tot), len(same), pct(len(same), len(tot))))

    # 이 표가 고객 안내문의 근거다. "이런 사진은 몇 % 실패한다" 를 숫자로 말할 수 있어야
    # 안내가 설득력을 갖는다. 유형은 겹칠 수 있으므로 건수 합계는 전체와 다르다.
    scored_rows = [r for r in rows if r.get("판정") != "오류"]
    if len(scored_rows) < len(rows):
        print("  ※ 아래 두 표는 판독에 성공한 %d건만 집계한다. "
              "오류 %d건은 판독 자체를 못 한 것이라 유형을 알 수 없다."
              % (len(scored_rows), len(rows) - len(scored_rows)))

    kinds = {}
    for r in scored_rows:
        for k in (r.get("이미지_문제") or ["(미분류)"]):
            kinds.setdefault(k, []).append(r)
    if kinds:
        print("[사진 유형별]  어떤 사진이 판독을 망치는가 — 고객 안내문의 근거")
        print("  %-12s %6s %8s %8s %8s" % ("유형", "건수", "성공", "불완전", "실패"))
        for name in sorted(kinds, key=lambda k: -len(kinds[k])):
            g = kinds[name]
            c = Counter(x.get("판정") for x in g)
            print("  %-12s %5d건 %8s %8s %8s"
                  % (name[:12], len(g), pct(c["성공"], len(g)),
                     pct(c["불완전"], len(g)), pct(c["실패"] + c["원본불량"], len(g))))
        print("  ※ 실패 열은 원본불량을 합친 값이다. 고객에게는 둘 다 '다시 보내주세요' 다")

    folders = {}
    for r in scored_rows:
        folders.setdefault(r.get("폴더") or "(최상위)", []).append(r)
    if len(folders) > 1:
        print("[폴더별]  묶음마다 성공률이 갈리면 그 묶음의 촬영·수집 방식에 원인이 있다")
        for name in sorted(folders):
            g = folders[name]
            c = Counter(x.get("판정") for x in g)
            print("  %-22s %4d건  성공 %s · 불완전 %s · 실패 %s"
                  % (name[:22], len(g), pct(c["성공"], len(g)),
                     pct(c["불완전"], len(g)), pct(c["실패"], len(g))))

    # 한도에 걸려 중간에 모델을 바꾸는 일이 생긴다. 그걸 한 숫자로 뭉치면
    # "정확도 85%" 가 어느 모델의 값인지 알 수 없어 판단 근거로 못 쓴다.
    models = {}
    for r in scored_rows:
        models.setdefault(r.get("모델") or "(미상)", []).append(r)
    if len(models) > 1:
        print("[모델별]  중간에 모델이 바뀌었다. 섞어서 보면 안 된다")
        print("  %-24s %6s %8s %8s %8s" % ("모델", "건수", "성공", "불완전", "실패"))
        for name in sorted(models, key=lambda k: -len(models[k])):
            g = models[name]
            c = Counter(x.get("판정") for x in g)
            print("  %-24s %5d건 %8s %8s %8s"
                  % (name[:24], len(g), pct(c["성공"], len(g)),
                     pct(c["불완전"], len(g)), pct(c["실패"] + c["원본불량"], len(g))))
        cross = {}
        for r in scored_rows:
            if r.get("우편번호일치"):
                cross.setdefault(r.get("모델") or "(미상)", []).append(r)
        if cross:
            print("  교차검증(사진에 우편번호가 찍힌 건)만 따로:")
            for name in sorted(cross):
                g = cross[name]
                hit = sum(1 for x in g if x["우편번호일치"] == "일치")
                print("    %-22s %d건 중 %d건 일치  %s" % (name[:22], len(g), hit, pct(hit, len(g))))

    lat = [float(r.get("지연초") or 0) for r in rows]
    tin = sum(int(r.get("입력토큰") or 0) for r in rows)
    tout = sum(int(r.get("출력토큰") or 0) for r in rows)
    usd = sum(LLM.cost_usd(r.get("모델"), int(r.get("입력토큰") or 0),
                           int(r.get("출력토큰") or 0)) for r in rows)
    print("[비용·시간]")
    print("  평균 %.1f초 · 최대 %.1f초 · 토큰 %s (입력 %s / 출력 %s)"
          % (sum(lat) / len(lat), max(lat), format(tin + tout, ","),
             format(tin, ","), format(tout, ",")))
    print("  총 %.3f USD ≈ %s원 (환율 %d 가정)"
          % (usd, format(int(round(usd * USD_KRW)), ","), USD_KRW))
    print("=" * 62)


# ------------------------------------------------------------------ 실행
def find_images(root, limit=0):
    out = []
    for dirpath, _, names in os.walk(root):
        for name in sorted(names):
            if os.path.splitext(name)[1].lower() in EXTS:
                out.append(os.path.join(dirpath, name))
    out.sort()
    return out[:limit] if limit else out


def sample_images(files, root, n, seed):
    """날짜 폴더 비율을 유지한 채 무작위로 n장을 뽑는다.

    13,000장을 전부 돌릴 이유가 없다. 재려는 것은 "이 챗봇의 판독 정확도"이지
    "이 13,000장 각각의 답"이 아니다. 표본이 수백 장을 넘어가면 오차범위가
    거의 줄지 않으므로, 그 뒤로는 시간과 돈만 더 든다.

    폴더별 비율을 맞추는 이유는 몇 달치를 모은 자료이기 때문이다. 무작위로만
    뽑으면 이미지가 많은 시기에 쏠려서, 그 시기의 촬영 습관이 전체 정확도인 것처럼
    나온다.

    같은 seed 면 같은 표본이 나온다. 프롬프트를 고친 뒤 같은 표본으로 다시 돌려야
    전후 비교가 성립한다."""
    if n <= 0 or n >= len(files):
        return files

    rnd = random.Random(seed)
    groups = {}
    for p in files:
        groups.setdefault(os.path.dirname(os.path.relpath(p, root)), []).append(p)

    picked, rest = [], []
    for key in sorted(groups):
        g = groups[key][:]
        rnd.shuffle(g)
        # 장수가 적은 폴더도 최소 한 장은 들어가게 한다
        take = max(1, int(round(n * len(g) / len(files))))
        picked.extend(g[:take])
        rest.extend(g[take:])

    # 폴더마다 반올림하면 합계가 요청한 수에 못 미친다.
    # 300장을 달라고 했는데 280장이 나오면 그 자체로 의심을 사므로 채워 넣는다.
    rnd.shuffle(rest)
    picked.extend(rest[:max(0, n - len(picked))])

    rnd.shuffle(picked)
    return sorted(picked[:n])


def main():
    ap = argparse.ArgumentParser(
        description="주소 이미지를 일괄 판독해 정확도를 측정한다")
    ap.add_argument("folder", help="이미지 폴더 (하위 폴더까지 훑는다)")
    ap.add_argument("--out", default="", help="결과 엑셀 경로 (기본 <outdir>/결과.xlsx)")
    ap.add_argument("--outdir", default="out", help="분류 폴더·진행 파일 위치 (기본 out)")
    ap.add_argument("--model", default="", help="모델명 (기본 secrets.toml MODELS 첫 항목)")
    ap.add_argument("--limit", type=int, default=0, help="앞에서 N장만")
    ap.add_argument("--sample", type=int, default=0,
                    help="무작위 표본 N장만 (날짜 폴더 비율 유지). 대량일 때 이걸 쓴다")
    ap.add_argument("--seed", type=int, default=42,
                    help="표본 고정값. 같은 값이면 같은 표본이 뽑힌다 (기본 42)")
    ap.add_argument("--workers", type=int, default=4, help="병렬 처리 수 (기본 4)")
    ap.add_argument("--repeat", type=int, default=1, help="같은 이미지를 N회 반복")
    ap.add_argument("--no-copy", action="store_true", help="판정별 폴더 복사 생략")
    args = ap.parse_args()

    if not os.path.isdir(args.folder):
        print("폴더가 없습니다: %s" % args.folder)
        return 1

    api_key = secret("GEMINI_API_KEY")
    if not api_key:
        print("GEMINI_API_KEY 를 찾지 못했습니다. .streamlit/secrets.toml 을 확인하세요.")
        return 1
    if not secret("JUSO_CONFM_KEY"):
        print("JUSO_CONFM_KEY 를 찾지 못했습니다. 주소 검증 없이는 채점이 불가능합니다.")
        return 1

    model = args.model or default_model()
    outdir = os.path.abspath(args.outdir)
    os.makedirs(outdir, exist_ok=True)
    xlsx = args.out or os.path.join(outdir, "결과.xlsx")
    store = Store(os.path.join(outdir, "진행.jsonl"))

    files = find_images(args.folder, args.limit)
    if not files:
        print("이미지가 없습니다: %s" % args.folder)
        return 1

    total_found = len(files)
    if args.sample:
        files = sample_images(files, args.folder, args.sample, args.seed)

    todo = [(p, run) for p in files for run in range(1, args.repeat + 1)
            if (os.path.relpath(p, args.folder), run) not in store.done]
    skipped = len(files) * args.repeat - len(todo)

    if args.sample and len(files) < total_found:
        print("전체 %d장 중 표본 %d장 (seed %d · 날짜 폴더 비율 유지)"
              % (total_found, len(files), args.seed))
        print("  같은 seed 로 다시 돌리면 같은 표본이 나옵니다. 프롬프트 수정 전후 비교에 쓰세요.")
    print("이미지 %d장 × %d회 = %d건" % (len(files), args.repeat, len(files) * args.repeat))
    if skipped:
        print("이미 처리된 %d건은 건너뜁니다 (%s)" % (skipped, store.path))
    print("모델 %s · 병렬 %d" % (model, args.workers))
    print("-" * 62)

    counter = {"n": 0}
    lock = threading.Lock()

    def work(item):
        path, run = item
        if _ERR["stop"]:
            # 남은 건은 건드리지 않는다. 지금 부르면 판독값은 못 쓰는데 돈만 나간다
            return None
        rec = process(path, args.folder, model, api_key, run)
        store.add(rec)
        with lock:
            counter["n"] += 1
            print("[%d/%d] %-8s %-6s %s %s"
                  % (counter["n"], len(todo), rec["판정"],
                     rec.get("우편번호") or "-", rec["파일명"],
                     ("· " + rec["오류"]) if rec["오류"] else ""), flush=True)
        return rec

    if todo:
        t0 = time.time()
        try:
            with ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
                list(ex.map(work, todo))
            print("-" * 62)
            if _ERR["stop"]:
                print("!" * 62)
                print("오류가 계속돼서 중단했습니다.")
                print("  사유: %s" % _ERR["why"])
                print("  하루 호출 한도라면 기다리는 것 말고 방법이 없습니다.")
                print("  다시 실행하면 성공한 건만 건너뛰고 나머지를 이어서 처리합니다.")
                print("!" * 62)
            else:
                print("%d건 처리 · %.1f분 소요" % (counter["n"], (time.time() - t0) / 60))
        except KeyboardInterrupt:
            # 여기까지 한 것은 이미 jsonl 에 있다. 그것으로 엑셀을 만들고 끝낸다
            print("\n중단했습니다. 처리한 %d건으로 결과를 만듭니다." % counter["n"])

    rows = mark_repeat(dedupe(store.rows))
    write_xlsx(rows, xlsx)
    print("엑셀: %s" % os.path.abspath(xlsx))

    if not args.no_copy:
        n = copy_by_verdict(rows, outdir)
        print("분류 복사: %s (새로 %d장) — 실패 폴더를 눈으로 훑으세요" % (outdir, n))

    summarize(rows)
    return 0


if __name__ == "__main__":
    sys.exit(main())
