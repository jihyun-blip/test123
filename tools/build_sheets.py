# -*- coding: utf-8 -*-
"""
기존 데모의 상품db.xlsx / 지침db.xlsx 를 설계서 3~4장 구조로 변환해
구글 시트에 import 할 CSV 11개를 생성한다.

  momo_master_products  : master_products
  momo_country_products : country_products, synonyms
  momo_bot_policies     : bot_policies
  momo_bot_logs         : conversations, turns, field_verdicts,
                          flag_verdicts, gaps, notes, policy_versions
"""
import csv
import io
import os
import sys
from collections import defaultdict

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SRC = r"C:\Users\Administrator\Desktop\모모플러스\챗봇\챗봇 테스트\momo_demo\config"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "sheets")

COUNTRY = "KR"
CURRENCY = "KRW"

# 품목코드 -> (축종, 부위, 비고). 마스터는 품목 정체성만 담당한다.
IDENTITY = {
    "A0013": ("돼지", "꼬리", ""),
    "B0023": ("돼지", "삼겹", ""),
    "A0022": ("돼지", "뒷다리", ""),
    "A0031": ("돼지", "뒷다리", "슬라이스 가공"),
    "A0024": ("돼지", "앞다리", ""),
    "A0026": ("소", "꼬리", ""),
}


def write_csv(spreadsheet, tab, rows):
    d = os.path.join(OUT, spreadsheet)
    os.makedirs(d, exist_ok=True)
    p = os.path.join(d, tab + ".csv")
    # utf-8-sig: 구글 시트 import 와 엑셀 더블클릭 양쪽에서 한글이 깨지지 않는다
    with open(p, "w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(rows)
    print("  %-22s %s행" % (tab + ".csv", len(rows) - 1))


# ---------------------------------------------------------------- 상품 3종
def build_products():
    wb = openpyxl.load_workbook(os.path.join(SRC, "상품db.xlsx"), data_only=True)
    ws = wb.worksheets[0]

    master = [["item_code", "canonical_name", "species", "part", "note"]]
    country = [["country_code", "item_code", "display_name",
                "price", "currency", "is_active", "note"]]
    syn = [["country_code", "item_code", "synonym", "note"]]

    owners = defaultdict(list)   # synonym -> [item_code]
    dropped_self = []

    for row in list(ws.iter_rows(values_only=True))[1:]:
        code, name, aliases, price = (str(c).strip() if c is not None else "" for c in row[:4])
        if not code:
            continue
        species, part, note = IDENTITY.get(code, ("", "", ""))
        master.append([code, name, species, part, note])
        country.append([COUNTRY, code, name, price, CURRENCY, "Y", ""])

        seen = set()
        for a in aliases.split(","):
            a = a.strip()
            if not a or a in seen:
                continue
            seen.add(a)
            # 자기 정식명과 같은 유사어는 버린다. 매칭 3단계(정식명 정확 일치)가 이미 잡는다
            if a == name:
                dropped_self.append((code, a))
                continue
            syn.append([COUNTRY, code, a, ""])
            owners[a].append(code)

    write_csv("momo_master_products", "master_products", master)
    write_csv("momo_country_products", "country_products", country)
    write_csv("momo_country_products", "synonyms", syn)

    canon = {r[0]: r[1] for r in master[1:]}
    return owners, canon, dropped_self


# ---------------------------------------------------------------- 지침
NEW_FLAGS = [
    ["플래그", "ADDRESS_IMAGE", "검수필수", "공통",
     "주소가 이미지에서 수집됨. 육안 확인 필요"],
    ["플래그", "ADDRESS_UNVERIFIED", "검수필수", "공통",
     "도로명주소 검색 API에서 우편번호 추출 실패"],
    ["플래그", "ADDRESS_DETAIL_MISSING", "미완료", "공통",
     "상세주소(동·호) 자체가 입력되지 않음"],
    ["플래그", "PHONE_MISMATCH", "미완료", "공통",
     "이미지에서 전화번호 2회 추출 결과 불일치"],
    ["플래그", "PRODUCT_SIGNAL_CONFLICT", "되물음", "공통",
     "라벨코드와 인쇄 상품명이 다른 상품을 지시"],
    ["플래그", "AMOUNT_MISMATCH", "차단", "공통",
     "챗봇 문장의 금액이 계산값과 불일치"],
]

# ADDRESS_DETAIL_MISSING 과 겹치지 않도록 기존 행의 설명을 좁힌다
DESC_FIX = {"ADDRESS_MISSING": "주소 자체가 없음 (상세주소 누락은 ADDRESS_DETAIL_MISSING)"}


def build_policies():
    wb = openpyxl.load_workbook(os.path.join(SRC, "지침db.xlsx"), data_only=True)
    ws = wb.worksheets[0]

    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row[:5]]
        if not any(cells):
            continue
        rows.append(cells)

    body = []
    fixed = []
    for r in rows[1:]:
        if r[1] in DESC_FIX:
            r = r[:4] + [DESC_FIX[r[1]]]
            fixed.append(r[1])
        body.append(r)

    existing = {r[1] for r in body}
    added = []
    for nf in NEW_FLAGS:
        if nf[1] in existing:
            continue
        body.append(nf)
        added.append(nf[1])

    write_csv("momo_bot_policies", "bot_policies", [rows[0]] + body)
    return fixed, added, len(body)


# ---------------------------------------------------------------- 로그 7종
LOGS = {
    "conversations": [
        "conversation_id", "started_at", "ended_at", "tester_name", "knowledge_mode",
        "model", "policy_version_id", "flag_settings_json", "turn_count", "image_count",
        "final_receiver", "final_phone", "final_address_base", "final_address_detail",
        "final_zipno", "final_road_addr", "item_count", "subtotal", "shipping_fee",
        "total", "app_version", "note",
    ],
    "turns": [
        "conversation_id", "turn_no", "timestamp", "user_text", "image_refs", "bot_text",
        "llm_raw_json", "state_diff_json", "used_refs_json", "missing_info_json",
        "flags_raised", "auto_detect_hits", "addr_keyword_raw", "addr_keyword_clean",
        "addr_error_code", "addr_total_count", "addr_road_addr", "addr_zipno",
        "latency_ms", "knowledge_mode", "policy_version_id", "tester_name",
    ],
    "field_verdicts": [
        "conversation_id", "field_type", "field_key", "final_value", "source",
        "source_ref", "source_turn", "verdict", "cause_tag", "tester_name",
        "knowledge_mode", "policy_version_id", "judged_at", "note",
    ],
    "flag_verdicts": [
        "conversation_id", "flag_key", "flag_value", "expected", "raised", "verdict",
        "behavior_verdict", "raised_turn", "evidence", "tester_name", "knowledge_mode",
        "policy_version_id", "judged_at", "note",
    ],
    "gaps": [
        "conversation_id", "turn_no", "asked", "needed", "found", "source_rule",
        "category", "tester_name", "knowledge_mode", "policy_version_id", "logged_at",
    ],
    "notes": [
        "conversation_id", "turn_no", "note_text", "tag", "tester_name",
        "knowledge_mode", "policy_version_id", "logged_at",
    ],
    "policy_versions": [
        "policy_version_id", "version_name", "memo", "base_version_id", "changed_keys",
        "diff_summary", "policies_json", "tester_name", "created_at",
    ],
}


def build_logs():
    for tab, header in LOGS.items():
        write_csv("momo_bot_logs", tab, [header])


# ---------------------------------------------------------------- 실행
print("[1] momo_master_products / momo_country_products")
owners, canon, dropped_self = build_products()

print("[2] momo_bot_policies")
fixed, added, n_policy = build_policies()

print("[3] momo_bot_logs")
build_logs()

print()
print("지침: %d행 (신규 %d행: %s / 설명수정: %s)"
      % (n_policy, len(added), ", ".join(added), ", ".join(fixed) or "없음"))
if dropped_self:
    print("자기 정식명과 동일해 제외한 유사어: "
          + ", ".join("%s(%s)" % (a, c) for c, a in dropped_self))

print()
print("--- 2개 이상 상품에 걸리는 유사어 (AMBIGUOUS_ALIAS 발동 지점) ---")
collide = {a: cs for a, cs in owners.items() if len(cs) > 1}
for a, cs in sorted(collide.items()):
    print("  %-12s -> %s" % (a, ", ".join("%s %s" % (c, canon[c]) for c in cs)))
print("  총 %d개" % len(collide))

print()
print("--- 정식명이면서 다른 상품의 유사어이기도 한 표현 (EXACT_NAME_PRIORITY 발동 지점) ---")
for code, name in canon.items():
    other = [c for c in owners.get(name, []) if c != code]
    if other:
        print("  %-12s -> 정식명 %s %s / 유사어 %s"
              % (name, code, name, ", ".join("%s %s" % (c, canon[c]) for c in other)))
