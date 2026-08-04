# -*- coding: utf-8 -*-
"""
build_sheets.py 가 만든 CSV 11개를, 탭 이름이 이미 박힌 xlsx 4개로 묶는다.
구글 드라이브에 업로드해 '구글 스프레드시트로 열기'만 하면 탭 구성이 그대로 된다.
"""
import csv
import io
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
SHEETS = os.path.join(BASE, "sheets")
OUT = os.path.join(BASE, "sheets_xlsx")

# 스프레드시트 파일 -> 탭 순서
BOOKS = {
    "momo_master_products":  ["master_products"],
    "momo_country_products": ["country_products", "synonyms"],
    "momo_bot_policies":     ["bot_policies"],
    "momo_bot_logs":         ["conversations", "turns", "field_verdicts",
                              "flag_verdicts", "gaps", "notes", "policy_versions"],
}

HEADER_FILL = PatternFill("solid", fgColor="E8EAED")
HEADER_FONT = Font(bold=True)

os.makedirs(OUT, exist_ok=True)

for book, tabs in BOOKS.items():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for tab in tabs:
        path = os.path.join(SHEETS, book, tab + ".csv")
        with open(path, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))

        ws = wb.create_sheet(title=tab)
        for r in rows:
            ws.append(r)

        # 1행 고정 + 강조. 어느 탭이든 헤더가 첫 행이라는 전제를 눈에 보이게 한다
        ws.freeze_panes = "A2"
        for c in range(1, len(rows[0]) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")

        # 열 너비를 내용에 맞춘다. 설명·JSON 컬럼이 길어 상한을 둔다
        for c in range(1, len(rows[0]) + 1):
            width = max((len(str(r[c - 1])) for r in rows if c - 1 < len(r)), default=10)
            ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 12), 46)

    dst = os.path.join(OUT, book + ".xlsx")
    wb.save(dst)
    print("%-24s 탭 %d개  %s" % (book + ".xlsx", len(tabs), " / ".join(tabs)))
